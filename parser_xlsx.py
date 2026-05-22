from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


# Campos que enriquecem cada Transacao além da chave de match (data+valor).
# Para planilha e Domínio são TODOS obrigatórios; para OFX nunca são preenchidos.
# data_pagamento é exclusivo da planilha (Domínio não tem esse campo direto).
CAMPOS_EXTRAS = (
    "data_emissao", "data_pagamento",
    "numero_nf", "cnpj", "fornecedor", "historico",
)


@dataclass
class Transacao:
    data: date                                 # vencimento (planilha/Domínio) ou data da transação (OFX)
    valor: Decimal
    descricao: str
    origem: str = "planilha"
    linha: int | None = None
    extras: dict[str, Any] = field(default_factory=dict)
    data_pagamento: date | None = None         # só planilha — usada no match com OFX


@dataclass
class EstruturaPlanilha:
    cabecalho: list[str]
    linhas: list[tuple] = field(default_factory=list)
    sugestao: dict[str, int] = field(default_factory=dict)
    linha_cabecalho: int = 1


# Aliases para auto-detecção das colunas obrigatórias (Data/Valor/Descrição)
DATA_ALIASES = {
    "data", "date", "dt", "dia",
    "data lancamento", "data lançamento", "data do lancamento", "data do lançamento",
    "data movimento", "data mov",
    "data pagamento", "data pgto", "data operacao", "data operação",
    "data credito", "data crédito", "data debito", "data débito",
    "vencimento", "data vencimento", "venc", "data venc",
    "data compensacao", "data compensação",
}
VALOR_ALIASES = {
    "valor", "value", "amount", "montante", "vlr", "vl",
    "total", "quantia", "r$", "valor (r$)", "valor r$",
    "valor lancamento", "valor lançamento", "valor parcela", "valor da parcela",
    "credito", "crédito", "debito", "débito",
    "entrada", "saida", "saída",
}
DESC_ALIASES = {
    "descricao", "descrição", "description", "memo", "obs", "observacao", "observação",
    "historico", "histórico", "detalhes", "detalhe",
    "lancamento", "lançamento", "evento", "transacao", "transação",
    "referencia", "referência", "ref",
    "documento", "doc", "complemento",
}

# Aliases para os campos extras opcionais
DATA_EMISSAO_ALIASES = {
    "data emissao", "data emissão", "emissao", "emissão",
    "data nf", "data da nota", "data documento", "dt emissao", "dt emissão",
}
DATA_PAGAMENTO_ALIASES = {
    "data pagamento", "data pagto", "data pgto", "pagamento", "pago em",
    "data de pagamento", "dt pgto", "dt pagto", "dt pagamento",
    "data baixa", "data quitacao", "data quitação", "data liquidacao",
    "data liquidação", "data compensacao", "data compensação",
}
NUMERO_NF_ALIASES = {
    "nf", "n nf", "n° nf", "nº nf", "numero nf", "número nf",
    "numero da nf", "número da nf", "num nf", "n nota", "nº nota",
    "numero", "número", "numero documento", "número documento", "doc",
    "nota", "nota fiscal", "numero nota fiscal",
}
CNPJ_ALIASES = {
    "cnpj", "cnpj fornecedor", "cnpj do fornecedor", "cgc", "cnpj/cpf",
    "documento fornecedor", "doc fornecedor",
}
FORNECEDOR_ALIASES = {
    "fornecedor", "nome fornecedor", "nome do fornecedor", "razao social",
    "razão social", "razao", "razão", "beneficiario", "beneficiário",
    "favorecido", "credor", "pagamento a", "para", "destinatario", "destinatário",
}
HISTORICO_ALIASES = {
    "historico", "histórico", "hist", "historico contabil", "histórico contábil",
    "descricao", "descrição", "complemento", "obs", "observacao", "observação",
    "detalhes", "detalhe", "narrativa",
}

ALIAS_MAP = {
    "data": DATA_ALIASES,
    "valor": VALOR_ALIASES,
    "data_emissao": DATA_EMISSAO_ALIASES,
    "data_pagamento": DATA_PAGAMENTO_ALIASES,
    "numero_nf": NUMERO_NF_ALIASES,
    "cnpj": CNPJ_ALIASES,
    "fornecedor": FORNECEDOR_ALIASES,
    "historico": HISTORICO_ALIASES,
}

# Apenas data (vencimento) e valor são chave de match — restante é opcional.
# Planilhas reais variam muito: algumas têm só vencimento+valor; outras
# trazem NF, CNPJ, fornecedor, histórico, etc.
CAMPOS_OBRIGATORIOS = ("data", "valor")


def _normaliza(texto: object) -> str:
    return str(texto).strip().lower() if texto is not None else ""


def para_data(valor: object) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


def para_decimal(valor: object) -> Decimal | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip()
    if not texto:
        return None
    negativo = False
    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1]
    texto = texto.replace("R$", "").replace("r$", "").replace(" ", "").replace("\xa0", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        d = Decimal(texto)
    except InvalidOperation:
        return None
    return -d if negativo else d


def _detecta_linha_cabecalho(linhas: list[tuple], max_busca: int = 15) -> int:
    """Procura a linha cujos rótulos batem com mais aliases. Empate → primeira."""
    melhor_idx, melhor_pontos = 0, -1
    for idx, linha in enumerate(linhas[:max_busca]):
        rotulos = [_normaliza(c) for c in linha]
        if sum(1 for r in rotulos if r) < 2:
            continue
        pontos = 0
        for r in rotulos:
            for aliases in ALIAS_MAP.values():
                if r in aliases:
                    pontos += 2
                    break
            else:
                if r and len(r) <= 40:
                    pontos += 1
        # Bônus: a próxima linha precisa ter dados
        if idx + 1 < len(linhas):
            prox = linhas[idx + 1]
            if any(c is not None and str(c).strip() for c in prox):
                pontos += 1
        if pontos > melhor_pontos:
            melhor_pontos, melhor_idx = pontos, idx
    return melhor_idx


def _mapeia_por_nome(cabecalho: list[str]) -> dict[str, int]:
    """Auto-detecta colunas obrigatórias E extras pelo nome do cabeçalho.

    Estratégia: prioriza aliases mais específicos primeiro (data_emissao tem
    'data emissao' que é mais específico que só 'data'; vencimento idem).
    """
    mapa: dict[str, int] = {}
    usados: set[int] = set()
    # Ordem importa: campos mais específicos antes dos genéricos
    # (histórico antes de fornecedor pra ele pegar "descricao"/"obs" se houver)
    ordem = (
        "data_pagamento", "data_emissao",
        "numero_nf", "cnpj", "historico", "fornecedor",
        "data", "valor",
    )
    for campo in ordem:
        aliases = ALIAS_MAP[campo]
        for idx, cel in enumerate(cabecalho):
            if idx in usados:
                continue
            if _normaliza(cel) in aliases:
                mapa[campo] = idx
                usados.add(idx)
                break
    return mapa


def _mapeia_por_conteudo(
    linhas_dados: list[tuple],
    n_colunas: int,
    ja_mapeado: dict[str, int],
    amostra: int = 30,
) -> dict[str, int]:
    """Quando o nome não bateu, classifica cada coluna pelo tipo do conteúdo."""
    usados = set(ja_mapeado.values())
    amostra_linhas = [l for l in linhas_dados[:amostra] if l]
    if not amostra_linhas:
        return {}

    scores_data = [0] * n_colunas
    scores_valor = [0] * n_colunas
    soma_len_texto = [0] * n_colunas
    n_strings = [0] * n_colunas

    for linha in amostra_linhas:
        for c in range(n_colunas):
            if c >= len(linha):
                continue
            cel = linha[c]
            if cel is None or cel == "":
                continue
            if para_data(cel) is not None and not isinstance(cel, (int, float)):
                scores_data[c] += 1
            if para_decimal(cel) is not None and not isinstance(cel, (datetime, date)):
                scores_valor[c] += 1
            if isinstance(cel, str) and cel.strip():
                soma_len_texto[c] += len(cel.strip())
                n_strings[c] += 1

    min_match = max(2, len(amostra_linhas) // 3)
    novo: dict[str, int] = {}

    if "data" not in ja_mapeado:
        candidatos = [(s, c) for c, s in enumerate(scores_data) if c not in usados and s >= min_match]
        if candidatos:
            idx = max(candidatos)[1]
            novo["data"] = idx
            usados.add(idx)

    if "valor" not in ja_mapeado:
        candidatos = [
            (s, c) for c, s in enumerate(scores_valor)
            if c not in usados and s >= min_match
        ]
        if candidatos:
            idx = max(candidatos)[1]
            novo["valor"] = idx
            usados.add(idx)

    # fornecedor por conteúdo: coluna com texto mais longo entre as não usadas
    if "fornecedor" not in ja_mapeado:
        candidatos = []
        for c in range(n_colunas):
            if c in usados or n_strings[c] == 0:
                continue
            media = soma_len_texto[c] / n_strings[c]
            if media >= 5:  # nome de fornecedor tem pelo menos 5 chars em média
                candidatos.append((media, c))
        if candidatos:
            idx = max(candidatos)[1]
            novo["fornecedor"] = idx

    return novo


def descobrir_estrutura(caminho: str | Path) -> EstruturaPlanilha:
    wb = load_workbook(filename=str(caminho), data_only=True, read_only=True)
    ws = wb.active
    todas = [tuple(row) for row in ws.iter_rows(values_only=True)]
    if not todas:
        return EstruturaPlanilha(cabecalho=[])

    idx_cab = _detecta_linha_cabecalho(todas)
    cab_raw = todas[idx_cab]
    cabecalho = [str(c).strip() if c is not None else "" for c in cab_raw]
    linhas_dados = todas[idx_cab + 1:]

    mapa = _mapeia_por_nome(cabecalho)
    # Se não detectou tudo pelo nome, tenta complementar pelo conteúdo
    if len(set(mapa.keys()) & set(CAMPOS_OBRIGATORIOS)) < len(CAMPOS_OBRIGATORIOS):
        mapa.update(_mapeia_por_conteudo(linhas_dados, len(cabecalho), mapa))

    return EstruturaPlanilha(
        cabecalho=cabecalho,
        linhas=linhas_dados,
        sugestao=mapa,
        linha_cabecalho=idx_cab + 1,
    )


def extrair_transacoes(estrutura: EstruturaPlanilha, mapeamento: dict[str, int]) -> list[Transacao]:
    faltando = set(CAMPOS_OBRIGATORIOS) - mapeamento.keys()
    if faltando:
        raise ValueError(f"Mapeamento incompleto: {', '.join(sorted(faltando))}")

    i_data = mapeamento["data"]
    i_valor = mapeamento["valor"]
    indices_extras = {
        campo: mapeamento[campo]
        for campo in CAMPOS_EXTRAS
        if campo in mapeamento and isinstance(mapeamento[campo], int) and mapeamento[campo] >= 0
    }
    base = estrutura.linha_cabecalho + 1

    transacoes: list[Transacao] = []
    for offset, linha in enumerate(estrutura.linhas):
        if not linha or all(c is None or c == "" for c in linha):
            continue
        data = para_data(linha[i_data]) if i_data < len(linha) else None
        valor = para_decimal(linha[i_valor]) if i_valor < len(linha) else None
        if data is None or valor is None:
            continue

        extras: dict[str, Any] = {}
        data_pagamento: date | None = None
        for campo, idx in indices_extras.items():
            if idx >= len(linha):
                continue
            valor_celula = linha[idx]
            if valor_celula is None or valor_celula == "":
                continue
            if campo == "data_emissao":
                d = para_data(valor_celula)
                if d is not None:
                    extras[campo] = d
            elif campo == "data_pagamento":
                d = para_data(valor_celula)
                if d is not None:
                    data_pagamento = d
                    extras[campo] = d
            else:
                extras[campo] = str(valor_celula).strip()

        transacoes.append(Transacao(
            data=data, valor=valor, descricao="",
            linha=base + offset, extras=extras,
            data_pagamento=data_pagamento,
        ))
    return transacoes


def ler_planilha(caminho: str | Path) -> list[Transacao]:
    """Atalho: descobre estrutura e extrai usando a sugestão automática.

    Levanta ValueError se a auto-detecção não cobrir as 3 colunas — nesse caso
    o chamador deve abrir o diálogo de mapeamento manual.
    """
    estrutura = descobrir_estrutura(caminho)
    return extrair_transacoes(estrutura, estrutura.sugestao)
